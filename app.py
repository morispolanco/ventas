import os
from dataclasses import dataclass, asdict
from io import BytesIO
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================================================
# SALES INTELLIGENCE GT
# Streamlit Cloud
# - Sin datos simulados
# - Sin créditos
# - Sin pagos
# - Sin autenticación
# - SerpAPI como fuente única
# - Búsqueda orientada a COMPRADORES potenciales
# - Selección múltiple y exportación a Excel formateado
# =========================================================

st.set_page_config(
    page_title="Sales Intelligence Guatemala",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
        .block-container { padding-top: 1.1rem; }
        .card {
            background: #ffffff;
            border: 1px solid #e5e7eb;
            border-radius: 16px;
            padding: 1rem 1rem 0.9rem 1rem;
            box-shadow: 0 1px 2px rgba(0,0,0,.04);
            margin-bottom: 0.9rem;
        }
        .badge {
            display: inline-block;
            padding: 0.25rem 0.65rem;
            border-radius: 999px;
            font-size: 0.8rem;
            font-weight: 600;
            margin-right: 0.4rem;
            margin-bottom: 0.25rem;
            border: 1px solid #e5e7eb;
            background: #f9fafb;
        }
        .badge-high { background: #ecfdf5; border-color: #a7f3d0; color: #065f46; }
        .badge-med { background: #fffbeb; border-color: #fde68a; color: #92400e; }
        .badge-low { background: #f3f4f6; border-color: #d1d5db; color: #374151; }
        .muted { color: #6b7280; font-size: 0.93rem; }
    </style>
    """,
    unsafe_allow_html=True,
)


@dataclass
class Prospect:
    name: str
    sector: str = ""
    location: str = ""
    size: str = ""
    website: str = ""
    source: str = ""
    signal: str = ""
    rationale: str = ""
    pain_point: str = ""
    score: int = 0
    probability: int = 0
    technologies: Optional[List[str]] = None
    evidence: Optional[List[str]] = None
    stage: str = "Nuevo"
    next_step: str = ""
    selected: bool = False

    def __post_init__(self) -> None:
        if self.technologies is None:
            self.technologies = []
        if self.evidence is None:
            self.evidence = []


def env(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


def split_csv(text: str) -> List[str]:
    return [part.strip() for part in text.split(",") if part.strip()]


def build_buyer_intent_query(product: str, sector: str, locations: List[str], signals: List[str]) -> str:
    buyer_terms = [
        "cotización",
        "cotizar",
        "busca proveedor",
        "buscando proveedor",
        "necesita",
        "requiere",
        "solicita",
        "implementación",
        "servicio",
        "comprar",
        "adquisición",
        "pedido",
        "licitación",
        "rfp",
        "rfi",
    ]

    exclude_terms = [
        "vender",
        "venta",
        "distribuidor",
        "mayorista",
        "catálogo",
        "tienda",
        "shop",
        "marketplace",
        "proveedor",
        "fabricante",
    ]

    query_parts: List[str] = []
    if product:
        query_parts.append(f'"{product}"')
    if sector:
        query_parts.append(sector)
    query_parts.extend(locations)
    query_parts.extend(signals)

    base = " ".join([p for p in query_parts if p]).strip()
    intent_block = " OR ".join([f'"{t}"' for t in buyer_terms])
    negative_block = " ".join([f'-{t}' for t in exclude_terms])

    return f"{base} ({intent_block}) {negative_block}".strip()


def serpapi_search(query: str, engine: str = "google", num: int = 10) -> List[Dict[str, Any]]:
    api_key = env("SERPAPI_KEY")
    if not api_key or not query.strip():
        return []

    url = "https://serpapi.com/search.json"
    params: Dict[str, Any] = {
        "engine": engine,
        "q": query,
        "api_key": api_key,
        "hl": "es",
        "gl": "gt",
        "num": min(max(num, 1), 10),
    }

    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
    except Exception:
        return []

    results: List[Dict[str, Any]] = []

    if engine == "google":
        for item in data.get("organic_results", [])[:num]:
            results.append(
                {
                    "title": item.get("title", ""),
                    "link": item.get("link", ""),
                    "snippet": item.get("snippet", ""),
                    "source": "SerpAPI Google",
                }
            )

        for item in data.get("news_results", [])[: max(0, num // 2)]:
            results.append(
                {
                    "title": item.get("title", ""),
                    "link": item.get("link", ""),
                    "snippet": item.get("snippet", "") or item.get("source", ""),
                    "source": "SerpAPI News",
                }
            )

    elif engine == "google_maps":
        for item in data.get("local_results", [])[:num]:
            results.append(
                {
                    "title": item.get("title", ""),
                    "link": item.get("website", "") or item.get("link", ""),
                    "snippet": item.get("address", "") or item.get("phone", ""),
                    "source": "SerpAPI Maps",
                }
            )

    return results


def build_prospect_from_result(result: Dict[str, Any], query: Dict[str, Any]) -> Prospect:
    title = (result.get("title") or "").strip() or "Resultado sin título"
    link = (result.get("link") or "").strip()
    snippet = (result.get("snippet") or "").strip()
    source = (result.get("source") or "Fuente real").strip()

    sector = query.get("buyer_sector", "")
    location = ", ".join(query.get("target_locations", [])) if query.get("target_locations") else ""

    rationale = (
        "Aparece en una consulta real de SerpAPI y contiene señales compatibles con intención de compra "
        "relacionadas con lo que el usuario vende."
    )
    pain_point = "Por confirmar con más evidencia en el sitio o fuente recuperada."

    return Prospect(
        name=title,
        sector=sector,
        location=location,
        size=query.get("size", ""),
        website=link,
        source=source,
        signal=snippet,
        rationale=rationale,
        pain_point=pain_point,
        evidence=[snippet] if snippet else [],
    )


def score_prospect(prospect: Dict[str, Any], query: Dict[str, Any]) -> int:
    score = 0
    sector = (prospect.get("sector") or "").lower()
    location = (prospect.get("location") or "").lower()
    signal = (prospect.get("signal") or "").lower()
    rationale = (prospect.get("rationale") or "").lower()
    evidence = prospect.get("evidence") or []

    target_sectors = [s.lower() for s in query.get("target_sectors", [])]
    target_locations = [s.lower() for s in query.get("target_locations", [])]
    signals = [s.lower() for s in query.get("signals", [])]

    if any(ts and ts in sector for ts in target_sectors):
        score += 20
    if any(tl and tl in location for tl in target_locations):
        score += 20
    if any(s and (s in signal or s in rationale) for s in signals):
        score += 30
    if prospect.get("website"):
        score += 10
    score += min(len(evidence) * 3, 15)

    return min(score, 100)


def create_formatted_excel(selected_rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospectos Seleccionados"

    title_fill = PatternFill("solid", fgColor="1F2937")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=14)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Nombre",
        "Sector",
        "Ubicación",
        "Tamaño",
        "Website",
        "Fuente",
        "Señal",
        "Razón",
        "Dolor probable",
        "Score",
        "Probabilidad de compra (%)",
        "Etapa",
        "Siguiente paso",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Prospectos seleccionados")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(selected_rows, start=3):
        values = [
            row.get("name", ""),
            row.get("sector", ""),
            row.get("location", ""),
            row.get("size", ""),
            row.get("website", ""),
            row.get("source", ""),
            row.get("signal", ""),
            row.get("rationale", ""),
            row.get("pain_point", ""),
            row.get("score", 0),
            row.get("probability", 0),
            row.get("stage", ""),
            row.get("next_step", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 30,
        2: 18,
        3: 18,
        4: 14,
        5: 38,
        6: 18,
        7: 42,
        8: 42,
        9: 28,
        10: 10,
        11: 20,
        12: 16,
        13: 20,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:M{len(selected_rows) + 2}"

    for row_idx in range(3, len(selected_rows) + 3):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="F9FAFB")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


if "pipeline" not in st.session_state:
    st.session_state.pipeline = []

if "selected_prospect" not in st.session_state:
    st.session_state.selected_prospect = None

if "search_results_df" not in st.session_state:
    st.session_state.search_results_df = pd.DataFrame()


with st.sidebar:
    st.title("💼 Sales Intelligence GT")
    st.caption("Inteligencia comercial real para Streamlit Cloud")
    st.markdown("---")

    st.subheader("Perfil de búsqueda")
    with st.form("search_form", clear_on_submit=False):
        product = st.text_input("¿Qué deseas vender?", placeholder="Ej. software ERP, impresión digital, consultoría, etc.")
        buyer_sector = st.text_input("Sector del comprador", placeholder="Ej. logística, retail, manufactura")
        target_locations_raw = st.text_input("Cobertura geográfica", placeholder="Ej. Ciudad de Guatemala, Escuintla")
        size = st.text_input("Tamaño del comprador", placeholder="Ej. 50-400 empleados")
        signals_raw = st.text_input("Señales de compra / intención", placeholder="Ej. cotización, busca proveedor, expansión, nueva sucursal")
        limit = st.slider("Máximo de resultados por fuente", 5, 20, 10)
        search_mode = st.selectbox("Modo de búsqueda", ["Web", "Mapas", "Ambos"])
        submitted = st.form_submit_button("Buscar compradores potenciales")

    show_over_80 = st.checkbox("Mostrar solo prospectos con probabilidad >= 80%", value=False)
    st.markdown("---")
    st.subheader("Exportación")
    st.caption("Solo se exportan los prospectos seleccionados.")


st.title("Cabina de Inteligencia Comercial")
st.write(
    "La aplicación busca y prioriza prospectos reales que muestran intención de compra de lo que tú vendes. "
    "No usa mock data, créditos, pagos ni autenticación."
)

if submitted:
    target_locations = split_csv(target_locations_raw)
    signals = split_csv(signals_raw)

    query_cfg: Dict[str, Any] = {
        "product": product,
        "buyer_sector": buyer_sector,
        "target_sectors": [buyer_sector] if buyer_sector else [],
        "target_locations": target_locations,
        "size": size,
        "signals": signals,
    }

    search_query = build_buyer_intent_query(product, buyer_sector, target_locations, signals)

    if not search_query.strip():
        st.warning("Completa al menos el producto o un criterio de búsqueda.")
    else:
        with st.spinner("Consultando SerpAPI..."):
            combined: List[Dict[str, Any]] = []
            if search_mode in ["Web", "Ambos"]:
                combined.extend(serpapi_search(search_query, engine="google", num=limit))
            if search_mode in ["Mapas", "Ambos"]:
                combined.extend(serpapi_search(search_query, engine="google_maps", num=limit))

            prospects: List[Dict[str, Any]] = []
            for result in combined:
                p = build_prospect_from_result(result, query_cfg)
                p.score = score_prospect(asdict(p), query_cfg)
                p.probability = p.score
                prospects.append(asdict(p))

            prospects = sorted(prospects, key=lambda x: x["score"], reverse=True)
            st.session_state.pipeline = prospects
            st.session_state.selected_prospect = prospects[0] if prospects else None

tab1, tab2, tab3 = st.tabs(["Compradores potenciales", "Pipeline", "Mensajes"])


with tab1:
    st.subheader("Compradores potenciales encontrados")
    st.caption("El resultado busca empresas que podrían querer comprar lo que tú vendes, no proveedores de la misma categoría.")

    pipeline = st.session_state.pipeline
    if show_over_80:
        pipeline = [p for p in pipeline if int(p.get("probability", 0)) >= 80]

    if not pipeline:
        st.info("No hay prospectos para mostrar con los filtros actuales.")
        st.markdown(
            """
            <div class="card">
                <strong>Estado vacío honesto</strong><br>
                Ajusta la búsqueda o desactiva el filtro de probabilidad mínima.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        df = pd.DataFrame(pipeline).copy()
        if "selected" not in df.columns:
            df["selected"] = False

        if st.checkbox("Seleccionar todos los visibles", key="select_all_visible"):
            df["selected"] = True

        st.caption("Marca los prospectos que deseas exportar.")
        edited_df = st.data_editor(
            df[
                [
                    "selected",
                    "name",
                    "probability",
                    "score",
                    "source",
                    "sector",
                    "location",
                    "website",
                    "signal",
                    "rationale",
                    "pain_point",
                    "stage",
                    "next_step",
                ]
            ],
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                "selected": st.column_config.CheckboxColumn("Seleccionar"),
                "probability": st.column_config.NumberColumn("Probabilidad (%)", min_value=0, max_value=100),
                "score": st.column_config.NumberColumn("Score", min_value=0, max_value=100),
                "website": st.column_config.LinkColumn("Website"),
            },
        )

        selected_rows = edited_df[edited_df["selected"] == True].copy()
        selected_rows = selected_rows.drop(columns=["selected"], errors="ignore")

        if not selected_rows.empty:
            st.success(f"{len(selected_rows)} prospecto(s) seleccionado(s).")
            excel_bytes = create_formatted_excel(selected_rows.to_dict(orient="records"))
            st.download_button(
                label="Exportar seleccionados a Excel",
                data=excel_bytes,
                file_name="prospectos_seleccionados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("Selecciona uno o varios prospectos para exportarlos.")


with tab2:
    st.subheader("Pipeline de trabajo")
    st.caption("Seguimiento simple, sin CRM de pago, sin login y sin créditos. El pipeline refleja empresas con intención de compra.")

    pipeline_df = pd.DataFrame(st.session_state.pipeline)
    if pipeline_df.empty:
        st.info("Todavía no hay prospectos en el pipeline.")
    else:
        if "stage" not in pipeline_df.columns:
            pipeline_df["stage"] = "Nuevo"
        if "next_step" not in pipeline_df.columns:
            pipeline_df["next_step"] = ""
        if "probability" not in pipeline_df.columns:
            pipeline_df["probability"] = pipeline_df.get("score", 0)

        cols = [c for c in ["name", "score", "probability", "source", "stage", "next_step", "website"] if c in pipeline_df.columns]
        edited = st.data_editor(
            pipeline_df[cols],
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            column_config={
                "stage": st.column_config.SelectboxColumn(
                    "stage",
                    options=["Nuevo", "Contactado", "Interesado", "Reunión", "Ganado", "Descartado"],
                    required=True,
                ),
                "probability": st.column_config.NumberColumn("Probabilidad (%)", min_value=0, max_value=100),
                "website": st.column_config.LinkColumn("Website"),
            },
        )

        if st.button("Guardar cambios del pipeline"):
            st.session_state.pipeline = edited.to_dict(orient="records")
            st.success("Pipeline actualizado.")


with tab3:
    st.subheader("Mensajes para compradores potenciales")

    prospect_dict = st.session_state.selected_prospect
    if not prospect_dict and st.session_state.pipeline:
        prospect_dict = st.session_state.pipeline[0]

    if not prospect_dict:
        st.info("Selecciona un prospecto para generar un mensaje.")
    else:
        col1, col2 = st.columns([1, 2])
        with col1:
            channel = st.selectbox("Canal", ["Correo", "LinkedIn", "WhatsApp"])
            tone = st.selectbox("Tono", ["Profesional", "Directo", "Consultivo"])
            generate = st.button("Generar mensaje")

        with col2:
            st.markdown(f"**Prospecto activo:** {prospect_dict.get('name', '')}")
            st.caption(f"Probabilidad de compra: {int(prospect_dict.get('probability', 0))}%")
            if generate:
                name = prospect_dict.get("name", "la empresa")
                signal = prospect_dict.get("signal", "una señal reciente observada en fuente real")
                product_name = env("DEFAULT_PRODUCT", "tu solución")

                if channel == "Correo":
                    body = (
                        f"Hola, equipo de {name}:\n\n"
                        f"Revisé {signal[:220]}. Creemos que {product_name} podría ayudarles a resolver una necesidad concreta relacionada con ese contexto.\n\n"
                        f"¿Te parece si coordinamos 10 minutos para compartirte una propuesta breve?\n\n"
                        f"Saludos,\n"
                        f"[Tu nombre]"
                    )
                elif channel == "LinkedIn":
                    body = (
                        f"Hola, vi una señal reciente sobre {name} y me pareció una buena razón para conectar.\n"
                        f"Trabajo con {product_name} y creo que podría ser relevante para su contexto actual."
                    )
                else:
                    body = (
                        f"Hola, soy [Tu nombre]. Vi una señal reciente sobre {name} y quería compartirte una idea corta sobre {product_name}.\n"
                        f"¿Te puedo enviar 3 líneas por aquí?"
                    )

                st.text_area("Borrador", value=body, height=220)
                st.caption(f"Tono sugerido: {tone}")

st.markdown("---")
st.caption("Configura SERPAPI_KEY en Streamlit Cloud para obtener resultados reales.")
